from __future__ import annotations
from dataclasses import dataclass
from typing import Any
from openpyxl import load_workbook
import io

REQUIRED_COLS = ["id_matricula", "id_disciplina", "id_avaliacao", "id_professor", "valor"]

@dataclass
class GradeRow:
    id_matricula: int
    id_disciplina: int
    id_avaliacao: int
    id_professor: int
    valor: float


def _to_int(v: Any, col: str) -> int:
    if v is None or str(v).strip() == "":
        raise ValueError(f"'{col}' vazio")
    return int(v)


def _to_nota(v: Any) -> float:
    """
    Converte a nota e valida se está entre 0 e 10
    """
    if v is None or str(v).strip() == "":
        raise ValueError("nota vazia")

    if isinstance(v, str):
        v = v.replace(",", ".").strip()

    valor = float(v)

    if valor < 0 or valor > 10:
        raise ValueError("nota fora do intervalo permitido (0 a 10)")

    return valor


def parse_excel_bytes(xlsx_bytes: bytes) -> list[GradeRow]:
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active

    header = [
        str(c.value).strip() if c.value is not None else ""
        for c in next(ws.iter_rows(min_row=1, max_row=1))
    ]

    col_map = {name: idx for idx, name in enumerate(header)}

    missing = [c for c in REQUIRED_COLS if c not in col_map]
    if missing:
        raise ValueError(f"Colunas faltando no Excel: {missing}")

    rows: list[GradeRow] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        
        if all(cell.value is None or str(cell.value).strip() == "" for cell in row):
            continue

        try:
            id_matricula = _to_int(row[col_map["id_matricula"]].value, "id_matricula")
            id_disciplina = _to_int(row[col_map["id_disciplina"]].value, "id_disciplina")
            id_avaliacao = _to_int(row[col_map["id_avaliacao"]].value, "id_avaliacao")
            id_professor = _to_int(row[col_map["id_professor"]].value, "id_professor")

            
            valor = _to_nota(row[col_map["valor"]].value)

            rows.append(
                GradeRow(
                    id_matricula=id_matricula,
                    id_disciplina=id_disciplina,
                    id_avaliacao=id_avaliacao,
                    id_professor=id_professor,
                    valor=valor,
                )
            )

        except Exception as e:
            raise ValueError(f"Erro na linha {row_idx}: {e}") from e

    if not rows:
        raise ValueError("Nenhuma nota válida encontrada no Excel.")

    return rows
