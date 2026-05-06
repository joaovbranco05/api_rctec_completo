from __future__ import annotations
from dataclasses import dataclass
from typing import Any
from openpyxl import load_workbook
import io

REQUIRED_COLS = [
    "nome do aluno",
    "nome da turma",
    "nome do professor",
    "nome da materia",
    "nome da avaliacao",
    "nota",
]

@dataclass
class GradeRowNamed:
    aluno_nome: str
    turma_nome: str
    professor_nome: str
    materia_nome: str
    avaliacao_nome: str
    nota: str 

def _cell_str(v: Any) -> str:
    return "" if v is None else str(v).strip()

def parse_excel_bytes_named(xlsx_bytes: bytes) -> list[GradeRowNamed]:
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active

    header = [_cell_str(c.value).lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col_map = {name: idx for idx, name in enumerate(header)}

    missing = [c for c in REQUIRED_COLS if c not in col_map]
    if missing:
        raise ValueError(f"Colunas faltando no Excel: {missing}")

    rows: list[GradeRowNamed] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if all(_cell_str(cell.value) == "" for cell in row):
            continue

        aluno = _cell_str(row[col_map["nome do aluno"]].value)
        turma = _cell_str(row[col_map["nome da turma"]].value)
        prof  = _cell_str(row[col_map["nome do professor"]].value)
        mat   = _cell_str(row[col_map["nome da materia"]].value)
        ava   = _cell_str(row[col_map["nome da avaliacao"]].value)
        nota  = _cell_str(row[col_map["nota"]].value)

        if not aluno or not turma or not prof or not mat or not ava:
            raise ValueError(f"Linha {row_idx}: existe campo vazio (exceto nota).")

        rows.append(
            GradeRowNamed(
                aluno_nome=aluno,
                turma_nome=turma,
                professor_nome=prof,
                materia_nome=mat,
                avaliacao_nome=ava,
                nota=nota,
            )
        )

    if not rows:
        raise ValueError("Nenhuma linha válida encontrada no Excel.")
    return rows
